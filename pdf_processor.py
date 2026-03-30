"""
Procesador de documentos para convertir requerimientos a texto.
Soporta PDF, Office, texto e imágenes.
"""
from __future__ import annotations

import os
import posixpath
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import List, Tuple, Optional
import xml.etree.ElementTree as ET

import requests

# Intentar importar librerías de PDF
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from pdf2image import convert_from_path
    HAS_PDF2IMAGE = True
except ImportError:
    HAS_PDF2IMAGE = False

try:
    import pytesseract
    from PIL import Image
    # En Windows, el binario puede no estar en PATH si la sesión no se reinició tras la instalación.
    # Configurar la ruta conocida si el binario no se encuentra por shutil.which.
    if shutil.which("tesseract") is None:
        import sys
        _win_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if sys.platform == "win32" and Path(_win_path).exists():
            pytesseract.pytesseract.tesseract_cmd = _win_path
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SUPPORTED_IMAGE_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"
}
SUPPORTED_TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".tsv", ".json", ".xml", ".yaml", ".yml", ".log", ".ini", ".cfg", ".sql"
}
SUPPORTED_OFFICE_EXTENSIONS = {".pptx", ".docx", ".xlsx"}
SUPPORTED_LEGACY_OFFICE_EXTENSIONS = {".doc", ".ppt", ".xls", ".odt", ".odp", ".ods"}
SUPPORTED_INPUT_EXTENSIONS = {
    ".pdf",
    *SUPPORTED_IMAGE_EXTENSIONS,
    *SUPPORTED_TEXT_EXTENSIONS,
    *SUPPORTED_OFFICE_EXTENSIONS,
    *SUPPORTED_LEGACY_OFFICE_EXTENSIONS,
}


class PDFProcessor:
    """Procesa documentos y los convierte a texto estructurado por páginas/diapositivas."""

    def __init__(
        self,
        use_ocr: bool = False,
        ocr_provider: str = "auto",
        nanonets_api_key: Optional[str] = None,
        nanonets_model_id: Optional[str] = None,
    ):
        """
        Args:
            use_ocr: Si True, activa OCR para documentos escaneados e imágenes.
            ocr_provider: auto | tesseract | nanonets
            nanonets_api_key: API key de Nanonets (si no, usa variable de entorno)
            nanonets_model_id: Model ID de Nanonets (si no, usa variable de entorno)
        """
        self.use_ocr = use_ocr
        self.ocr_provider = (ocr_provider or "auto").strip().lower()
        self.nanonets_api_key = nanonets_api_key or os.getenv("NANONETS_API_KEY", "")
        self.nanonets_model_id = nanonets_model_id or os.getenv("NANONETS_MODEL_ID", "")

        if self.ocr_provider not in {"auto", "tesseract", "nanonets", "llm"}:
            raise ValueError("ocr_provider debe ser: auto, tesseract, nanonets o llm")

    def _has_nanonets_config(self) -> bool:
        return bool(self.nanonets_api_key and self.nanonets_model_id)

    def _resolve_tessdata_prefix(self) -> Optional[str]:
        """Detecta carpeta tessdata para tesseract."""
        def is_valid_tessdata_dir(path_str: str) -> bool:
            p = Path(path_str)
            if not p.exists() or not p.is_dir():
                return False
            return any(p.glob("*.traineddata"))

        env_prefix = os.getenv("TESSDATA_PREFIX", "").strip()
        if env_prefix and is_valid_tessdata_dir(env_prefix):
            return env_prefix

        candidates = [
            r"C:\Program Files\Tesseract-OCR\tessdata",
            "/opt/homebrew/share/tessdata",
            "/usr/local/share/tessdata",
            "/usr/share/tesseract-ocr/4.00/tessdata",
            "/usr/share/tesseract-ocr/5/tessdata",
            "/usr/share/tessdata",
        ]
        for c in candidates:
            if is_valid_tessdata_dir(c):
                return c
        return None

    def _nanonets_url(self) -> str:
        return f"https://app.nanonets.com/api/v2/OCR/Model/{self.nanonets_model_id}/LabelFile/"

    def _collect_strings(self, value) -> List[str]:
        """Extrae texto de respuesta OCR de Nanonets de forma defensiva."""
        out: List[str] = []

        if isinstance(value, str):
            txt = value.strip()
            if txt:
                out.append(txt)
            return out

        if isinstance(value, list):
            for item in value:
                out.extend(self._collect_strings(item))
            return out

        if isinstance(value, dict):
            priority_keys = ("ocr_text", "text", "label", "value")
            for key in priority_keys:
                if key in value:
                    out.extend(self._collect_strings(value[key]))
            for key, item in value.items():
                if key not in priority_keys:
                    out.extend(self._collect_strings(item))
            return out

        return out

    def extract_text_from_image_nanonets(self, image_path: Path) -> str:
        """Extrae texto de imagen usando Nanonets OCR."""
        if not self._has_nanonets_config():
            raise RuntimeError(
                "Falta configuración de Nanonets. "
                "Define NANONETS_API_KEY y NANONETS_MODEL_ID"
            )

        with open(image_path, "rb") as f:
            response = requests.post(
                self._nanonets_url(),
                auth=(self.nanonets_api_key, ""),
                files={"file": (image_path.name, f)},
                data={"modelId": self.nanonets_model_id},
                timeout=180,
            )

        response.raise_for_status()
        payload = response.json()
        candidates = self._collect_strings(payload)

        # Deduplicar conservando orden
        seen = set()
        ordered = []
        for txt in candidates:
            key = txt.lower()
            if key in seen:
                continue
            seen.add(key)
            ordered.append(txt)

        return "\n".join(ordered).strip()

    def extract_text_from_image_tesseract(self, image_path: Path) -> str:
        """Extrae texto de imagen usando Tesseract."""
        tessdata_dir = self._resolve_tessdata_prefix()

        # Usar TESSDATA_PREFIX como variable de entorno en lugar de --tessdata-dir
        # para evitar problemas con rutas que contienen espacios en Windows.
        if tessdata_dir:
            os.environ.setdefault("TESSDATA_PREFIX", tessdata_dir)

        if HAS_TESSERACT:
            with Image.open(image_path) as img:
                for lang in ("spa", "eng"):
                    try:
                        txt = pytesseract.image_to_string(img, lang=lang).strip()
                        if txt:
                            return txt
                    except Exception:
                        continue
                # Último intento sin idioma explícito
                return pytesseract.image_to_string(img).strip()

        # Fallback sin dependencias Python: usar tesseract CLI directamente.
        tesseract_bin = shutil.which("tesseract")
        if not tesseract_bin:
            raise ImportError(
                "OCR Tesseract no disponible. Instala pytesseract+Pillow "
                "o el binario 'tesseract'."
            )

        last_err = ""
        env = os.environ.copy()
        if tessdata_dir:
            env["TESSDATA_PREFIX"] = tessdata_dir

        for lang in ("spa", "eng"):
            cmd = [tesseract_bin, str(image_path), "stdout", "-l", lang]
            res = subprocess.run(cmd, capture_output=True, text=True, env=env)
            if res.returncode == 0 and (res.stdout or "").strip():
                return (res.stdout or "").strip()
            last_err = (res.stderr or res.stdout or "").strip()

        # Último intento sin idioma explícito.
        cmd = [tesseract_bin, str(image_path), "stdout"]
        res = subprocess.run(cmd, capture_output=True, text=True, env=env)
        if res.returncode == 0:
            return (res.stdout or "").strip()

        err = (res.stderr or res.stdout or last_err or "").strip()
        raise RuntimeError(f"tesseract CLI falló: {err}")

    def _get_ollama_base_url(self) -> str:
        """Devuelve la URL base de Ollama sin el sufijo /v1."""
        from config import config as llm_config
        base = llm_config.base_url.rstrip("/")
        if base.endswith("/v1"):
            base = base[:-3]
        return base

    def _find_ollama_vision_model(self) -> Optional[str]:
        """Busca un modelo con capacidad de visión en Ollama. Prefiere el modelo OCR."""
        preferred = "blaifa/Nanonets-OCR-s:3b-q8_0"
        try:
            resp = requests.get(f"{self._get_ollama_base_url()}/api/tags", timeout=10)
            resp.raise_for_status()
            models = [m["name"] for m in resp.json().get("models", [])]
        except Exception:
            return None

        if preferred in models:
            return preferred

        # Buscar cualquier modelo con capacidad vision
        for name in models:
            try:
                info = requests.post(
                    f"{self._get_ollama_base_url()}/api/show",
                    json={"name": name},
                    timeout=10,
                ).json()
                if "vision" in info.get("capabilities", []):
                    return name
            except Exception:
                continue
        return None

    def extract_text_from_image_llm(self, image_path: Path) -> str:
        """Extrae texto de imagen usando un modelo de visión local de Ollama (OCR via LLM)."""
        import base64

        model = self._find_ollama_vision_model()
        if not model:
            raise RuntimeError(
                "No se encontró modelo de visión en Ollama. "
                "Instala 'blaifa/Nanonets-OCR-s:3b-q8_0' con: ollama pull blaifa/Nanonets-OCR-s:3b-q8_0"
            )

        with open(image_path, "rb") as f:
            image_b64 = base64.b64encode(f.read()).decode("utf-8")

        prompt = (
            "Extract all text from this image exactly as it appears. "
            "Return only the extracted text, without explanations or formatting markers."
        )

        resp = requests.post(
            f"{self._get_ollama_base_url()}/api/generate",
            json={
                "model": model,
                "prompt": prompt,
                "images": [image_b64],
                "stream": False,
            },
            timeout=120,
        )
        resp.raise_for_status()
        return resp.json().get("response", "").strip()

    def extract_text_from_image(self, image_path: Path, force: bool = False) -> str:
        """Extrae texto de imagen según provider configurado."""
        if not self.use_ocr and not force:
            return ""

        provider = self.ocr_provider

        if provider == "nanonets":
            return self.extract_text_from_image_nanonets(image_path)

        if provider == "tesseract":
            return self.extract_text_from_image_tesseract(image_path)

        if provider == "llm":
            return self.extract_text_from_image_llm(image_path)

        # auto: prioriza Nanonets → Tesseract → LLM local.
        if self._has_nanonets_config():
            try:
                return self.extract_text_from_image_nanonets(image_path)
            except Exception as e:
                print(f"⚠️  Nanonets OCR falló en {image_path.name}: {e}. Fallback a Tesseract...")

        try:
            return self.extract_text_from_image_tesseract(image_path)
        except Exception as e:
            print(f"⚠️  Tesseract OCR falló en {image_path.name}: {e}. Fallback a LLM local...")

        return self.extract_text_from_image_llm(image_path)

    def extract_text_pypdf2(self, pdf_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto usando PyPDF2."""
        pages = []
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            total = len(reader.pages)

            for i, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                pages.append((i, total, text))

        return pages

    def extract_text_pdfplumber(self, pdf_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto usando pdfplumber (mejor para tablas)."""
        pages = []
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)

            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                pages.append((i, total, text))

        return pages

    def extract_text_ocr(self, pdf_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto de PDF escaneado usando OCR (Nanonets o Tesseract)."""
        if not HAS_PDF2IMAGE:
            raise ImportError(
                "OCR para PDF requiere pdf2image. "
                "Instala con: pip install pdf2image"
            )

        images = convert_from_path(pdf_path)
        total = len(images)
        pages = []

        with tempfile.TemporaryDirectory(prefix="ocr_pdf_") as tmp_dir:
            tmp_root = Path(tmp_dir)

            for i, image in enumerate(images, 1):
                text = ""

                if self.use_ocr:
                    tmp_image = tmp_root / f"{pdf_path.stem}_p{i}.png"
                    image.save(tmp_image, format="PNG")
                    text = self.extract_text_from_image(tmp_image)

                pages.append((i, total, text or ""))

        return pages

    def _extract_text_from_slide_xml(self, slide_xml: bytes) -> str:
        """Extrae texto visible de un slide XML de PPTX."""
        try:
            root = ET.fromstring(slide_xml)
        except ET.ParseError:
            return ""

        texts: List[str] = []
        for node in root.iter():
            if node.tag.endswith("}t") and node.text:
                txt = node.text.strip()
                if txt:
                    texts.append(txt)

        return "\n".join(texts).strip()

    def _extract_slide_media_paths(self, zf: zipfile.ZipFile, slide_path: str) -> List[str]:
        """Obtiene rutas de imágenes asociadas a un slide dentro del PPTX."""
        rels_path = posixpath.join(
            posixpath.dirname(slide_path),
            "_rels",
            f"{posixpath.basename(slide_path)}.rels",
        )

        if rels_path not in zf.namelist():
            return []

        try:
            rels_root = ET.fromstring(zf.read(rels_path))
        except ET.ParseError:
            return []

        media_paths: List[str] = []
        for rel in rels_root.iter():
            if not rel.tag.endswith("Relationship"):
                continue

            rel_type = str(rel.attrib.get("Type", ""))
            target = str(rel.attrib.get("Target", "")).strip()

            if not target or not rel_type.endswith("/image"):
                continue

            resolved = posixpath.normpath(
                posixpath.join(posixpath.dirname(slide_path), target)
            )
            if resolved in zf.namelist():
                media_paths.append(resolved)

        return media_paths

    def extract_text_pptx(self, pptx_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto de todas las diapositivas del PPTX en orden."""
        pages: List[Tuple[int, int, str]] = []

        with zipfile.ZipFile(pptx_path, "r") as zf:
            slide_paths = [
                n for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            ]

            def slide_num(name: str) -> int:
                stem = Path(name).stem  # slide12
                digits = "".join(ch for ch in stem if ch.isdigit())
                return int(digits) if digits else 10**9

            slide_paths.sort(key=slide_num)
            total = len(slide_paths)

            for idx, slide_path in enumerate(slide_paths, 1):
                text_direct = self._extract_text_from_slide_xml(zf.read(slide_path))
                slide_parts: List[str] = []

                if text_direct:
                    slide_parts.append(text_direct)

                if self.use_ocr:
                    media_paths = self._extract_slide_media_paths(zf, slide_path)
                    media_ocr_parts: List[str] = []

                    # Filtrar formatos vectoriales que Pillow/Tesseract no pueden leer
                    RASTER_ONLY = SUPPORTED_IMAGE_EXTENSIONS - {".svg", ".emf", ".wmf"}
                    media_paths = [m for m in media_paths if Path(m).suffix.lower() in RASTER_ONLY]

                    for media_path in media_paths:
                        suffix = Path(media_path).suffix or ".img"
                        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                            tmp.write(zf.read(media_path))
                            tmp_file = Path(tmp.name)

                        try:
                            ocr_text = self.extract_text_from_image(tmp_file)
                            if ocr_text.strip():
                                media_ocr_parts.append(ocr_text.strip())
                        finally:
                            try:
                                tmp_file.unlink(missing_ok=True)
                            except Exception:
                                pass

                    if media_ocr_parts:
                        slide_parts.append("[OCR IMÁGENES]\n" + "\n\n".join(media_ocr_parts))

                pages.append((idx, total, "\n\n".join(slide_parts).strip()))

        return pages

    def extract_text_docx(self, docx_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto de DOCX en orden de lectura aproximado."""
        with zipfile.ZipFile(docx_path, "r") as zf:
            if "word/document.xml" not in zf.namelist():
                return [(1, 1, "")]

            try:
                root = ET.fromstring(zf.read("word/document.xml"))
            except ET.ParseError:
                return [(1, 1, "")]

            chunks: List[str] = []
            for node in root.iter():
                if node.tag.endswith("}p"):
                    run_parts: List[str] = []
                    for sub in node.iter():
                        if sub.tag.endswith("}t") and sub.text:
                            txt = sub.text.strip()
                            if txt:
                                run_parts.append(txt)
                    if run_parts:
                        chunks.append(" ".join(run_parts))

            if self.use_ocr:
                media_paths = [
                    n for n in zf.namelist()
                    if n.startswith("word/media/") and Path(n).suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
                ]
                ocr_chunks: List[str] = []
                for media_path in media_paths:
                    suffix = Path(media_path).suffix or ".img"
                    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                        tmp.write(zf.read(media_path))
                        tmp_file = Path(tmp.name)
                    try:
                        txt = self.extract_text_from_image(tmp_file).strip()
                        if txt:
                            ocr_chunks.append(txt)
                    finally:
                        try:
                            tmp_file.unlink(missing_ok=True)
                        except Exception:
                            pass
                if ocr_chunks:
                    chunks.append("[OCR IMÁGENES DOCX]\n" + "\n\n".join(ocr_chunks))

            return [(1, 1, "\n".join(chunks).strip())]

    def extract_text_xlsx(self, xlsx_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto de XLSX por hoja."""
        if not HAS_OPENPYXL:
            raise ImportError("XLSX requiere openpyxl. Instala con: pip install openpyxl")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        pages: List[Tuple[int, int, str]] = []
        total = len(wb.sheetnames) if wb.sheetnames else 1

        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            lines = [f"[HOJA] {sheet_name}"]
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    lines.append(" | ".join(values))
            pages.append((idx, total, "\n".join(lines).strip()))

        wb.close()
        return pages or [(1, 1, "")]

    def extract_text_plain(self, file_path: Path) -> List[Tuple[int, int, str]]:
        """Extrae texto de archivos text/plain-like."""
        try:
            txt = file_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            txt = file_path.read_text(encoding="latin-1", errors="replace")
        return [(1, 1, txt)]

    def _convert_legacy_office_via_com(self, input_path: Path, out_dir: Path) -> Optional[Path]:
        """
        Convierte archivos Office legacy (.ppt/.doc/.xls) a PPTX/DOCX/XLSX
        usando la aplicación Office instalada en Windows (COM automation).
        Retorna la ruta del archivo convertido, o None si falla.
        """
        import sys
        if sys.platform != "win32":
            return None
        try:
            import win32com.client
        except ImportError:
            return None

        ext = input_path.suffix.lower()
        abs_input = str(input_path.resolve())

        try:
            if ext == ".ppt":
                app = win32com.client.Dispatch("PowerPoint.Application")
                prs = app.Presentations.Open(abs_input, WithWindow=False)
                out_path = out_dir / f"{input_path.stem}.pptx"
                # 24 = ppSaveAsOpenXMLPresentation
                prs.SaveAs(str(out_path.resolve()), 24)
                prs.Close()
                app.Quit()
                return out_path if out_path.exists() else None

            if ext == ".doc":
                app = win32com.client.Dispatch("Word.Application")
                doc = app.Documents.Open(abs_input)
                out_path = out_dir / f"{input_path.stem}.docx"
                # 16 = wdFormatXMLDocument
                doc.SaveAs2(str(out_path.resolve()), 16)
                doc.Close()
                app.Quit()
                return out_path if out_path.exists() else None

            if ext == ".xls":
                app = win32com.client.Dispatch("Excel.Application")
                wb = app.Workbooks.Open(abs_input)
                out_path = out_dir / f"{input_path.stem}.xlsx"
                # 51 = xlOpenXMLWorkbook
                wb.SaveAs(str(out_path.resolve()), 51)
                wb.Close()
                app.Quit()
                return out_path if out_path.exists() else None

        except Exception:
            pass
        return None

    def extract_text_legacy_office(self, input_path: Path, method: str = "auto") -> List[Tuple[int, int, str]]:
        """
        Convierte archivos office legacy (.ppt/.doc/.xls) a formato moderno y extrae texto.
        Prioridad: 1) COM de Office (Windows), 2) LibreOffice (soffice).
        """
        with tempfile.TemporaryDirectory(prefix="legacy_office_") as tmp_dir:
            out_dir = Path(tmp_dir)

            # Intento 1: COM de Office (Windows, sin dependencias extra)
            converted = self._convert_legacy_office_via_com(input_path, out_dir)
            if converted:
                return self.extract_text_document(converted, method=method)

            # Intento 2: LibreOffice
            soffice = shutil.which("soffice")
            if not soffice:
                raise RuntimeError(
                    f"{input_path.suffix} requiere LibreOffice para conversión automática. "
                    "Instala LibreOffice (comando 'soffice') o convierte el archivo a PDF/PPTX/DOCX/XLSX."
                )

            cmd = [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(input_path)]
            res = subprocess.run(cmd, capture_output=True, text=True)
            if res.returncode != 0:
                raise RuntimeError(
                    f"No se pudo convertir {input_path.name} a PDF con soffice: {res.stderr.strip() or res.stdout.strip()}"
                )

            pdf_path = out_dir / f"{input_path.stem}.pdf"
            if not pdf_path.exists():
                raise RuntimeError(f"Conversión a PDF incompleta para {input_path.name}")

            return self.extract_text(pdf_path, method=method)

    def extract_text_document(self, input_path: Path, method: str = "auto") -> List[Tuple[int, int, str]]:
        """Extrae texto según tipo de documento de entrada."""
        if not input_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {input_path}")

        ext = input_path.suffix.lower()
        print(f"Extrayendo texto de {input_path.name}...")

        if ext == ".pdf":
            return self.extract_text(input_path, method=method)

        if ext == ".pptx":
            return self.extract_text_pptx(input_path)

        if ext == ".docx":
            return self.extract_text_docx(input_path)

        if ext == ".xlsx":
            return self.extract_text_xlsx(input_path)

        if ext in SUPPORTED_TEXT_EXTENSIONS:
            return self.extract_text_plain(input_path)

        if ext in SUPPORTED_LEGACY_OFFICE_EXTENSIONS:
            return self.extract_text_legacy_office(input_path, method=method)

        if ext in SUPPORTED_IMAGE_EXTENSIONS:
            # Las imágenes requieren OCR sí o sí para extraer contenido.
            try:
                text = self.extract_text_from_image(input_path, force=True)
            except Exception as e:
                print(f"⚠️  OCR de imagen falló en {input_path.name}: {e}")
                text = ""
            return [(1, 1, text)]

        raise ValueError(f"Formato no soportado: {input_path.suffix}")

    def extract_text(self, pdf_path: Path, method: str = "auto") -> List[Tuple[int, int, str]]:
        """
        Extrae texto del PDF usando el método especificado.

        Args:
            pdf_path: Ruta al archivo PDF
            method: "auto", "pypdf2", "pdfplumber", "ocr"

        Returns:
            Lista de tuplas (page_num, total_pages, text)
        """
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF no encontrado: {pdf_path}")

        if method == "ocr" or (method == "auto" and self.use_ocr):
            return self.extract_text_ocr(pdf_path)

        # Intentar pdfplumber primero (mejor calidad)
        if method == "pdfplumber" or (method == "auto" and HAS_PDFPLUMBER):
            try:
                pages = self.extract_text_pdfplumber(pdf_path)
                # Verificar si extrajo algo útil
                total_chars = sum(len(text) for _, _, text in pages)
                if total_chars > 100:
                    return pages
            except Exception as e:
                print(f"⚠️  pdfplumber falló: {e}")

        # Fallback a PyPDF2
        if HAS_PYPDF2:
            return self.extract_text_pypdf2(pdf_path)

        raise RuntimeError(
            "No se pudo extraer texto del PDF. "
            "Instala PyPDF2/pdfplumber o usa --pdf-method ocr con --use-ocr."
        )

    def save_as_txt(self, input_path: Path, output_path: Path, method: str = "auto") -> Path:
        """
        Convierte documento a TXT con formato de páginas/diapositivas.

        Args:
            input_path: Ruta al documento
            output_path: Ruta para el TXT de salida
            method: Método de extracción para PDF

        Returns:
            Ruta al archivo TXT generado
        """
        pages = self.extract_text_document(input_path, method=method)

        with open(output_path, "w", encoding="utf-8") as f:
            for page_num, total_pages, text in pages:
                f.write(f"--- PÁGINA {page_num} / {total_pages} (TEXTO DIRECTO) ---\n")
                f.write(text)
                f.write("\n\n")

        print(f"  ✓ Guardado en {output_path}")
        return output_path

    def process_folder(
        self,
        folder: Path,
        output_folder: Path,
        method: str = "auto",
        pattern: str = "*",
        recursive: bool = True,
        pdf_files: Optional[List[Path]] = None,
        input_files: Optional[List[Path]] = None,
    ) -> List[Path]:
        """
        Procesa documentos de una carpeta.

        Args:
            folder: Carpeta con documentos
            output_folder: Carpeta para TXTs
            method: Método de extracción para PDF
            pattern: Patrón de búsqueda (si no hay lista explícita)
            recursive: Si True busca en subcarpetas
            pdf_files: Compatibilidad hacia atrás (lista explícita)
            input_files: Lista explícita de documentos a procesar

        Returns:
            Lista de rutas a los TXTs generados
        """
        output_folder.mkdir(parents=True, exist_ok=True)

        explicit_files = input_files if input_files is not None else pdf_files

        if explicit_files is not None:
            files_to_process = [
                p for p in explicit_files
                if p.exists() and p.is_file() and p.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS
            ]
        else:
            if recursive:
                candidates = list(folder.rglob(pattern))
            else:
                candidates = list(folder.glob(pattern))

            files_to_process = [
                p for p in candidates
                if p.is_file() and p.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS
            ]

        if not files_to_process:
            print(f"⚠️  No se encontraron documentos soportados en {folder}")
            return []

        print(f"📄 Procesando {len(files_to_process)} documento(s)...")
        txt_files = []

        for input_path in sorted(files_to_process):
            # Generar ruta de salida preservando la estructura relativa
            rel_parent = input_path.parent.relative_to(folder)
            target_dir = output_folder / rel_parent
            target_dir.mkdir(parents=True, exist_ok=True)

            txt_name = f"{input_path.stem}{input_path.suffix.lower().replace('.', '_')}.txt"
            txt_path = target_dir / txt_name

            try:
                self.save_as_txt(input_path, txt_path, method=method)
                txt_files.append(txt_path)
            except Exception as e:
                print(f"❌ Error procesando {input_path.name}: {e}")
                continue

        return txt_files


def convert_pdfs_to_txt(
    input_folder: Path,
    output_folder: Path,
    use_ocr: bool = False,
    method: str = "auto",
    recursive: bool = True,
) -> List[Path]:
    """
    Helper de compatibilidad para convertir documentos a TXT.
    """
    processor = PDFProcessor(use_ocr=use_ocr)
    return processor.process_folder(
        input_folder,
        output_folder,
        method=method,
        recursive=recursive,
    )


if __name__ == "__main__":
    # Prueba
    base_dir = Path(__file__).parent
    input_dir = base_dir / "requerimientos"
    output_dir = base_dir / "requerimientos" / "txt"

    if not input_dir.exists():
        print(f"Crear carpeta: {input_dir}")
        input_dir.mkdir(parents=True)

    processor = PDFProcessor()
    txt_files = processor.process_folder(input_dir, output_dir)
    print(f"\n✅ Procesados {len(txt_files)} archivo(s)")
