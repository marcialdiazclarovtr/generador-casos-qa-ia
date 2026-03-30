"""
Agente Maestro: Orquestador principal del sistema multi-agente.

Responsabilidades:
1. Planificar qué combinaciones de casos generar (analizando JSON + matriz)
2. Disparar Agente 1 (cabecera) y Agente 2 (detalle) secuencialmente
3. Validar outputs de cada agente contra las reglas de negocio
4. Acumular casos validados y exportar a CSV (formato Mantis)
"""
import csv
import json
import random
import re
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config import agent_config
from llm_client import LMStudioClient, JSONExtractor, get_llm_client
from knowledge_loader import KnowledgeBase, PROCESO_DIC_A_MATRIZ
from agente1_campos import Agente1Campos
from agente2_detalle import Agente2Detalle

# Empaquetados del diccionario
EMPAQUETADOS = ["1Play", "2Play", "3Play"]


# ── Columnas del CSV de salida (formato Mantis exacto) ──────────────────────

COLUMNAS_CSV = [
    "ID",
    "Tipo de Prueba",
    "Prioridad",
    "Marca",
    "Segmento",
    "Tecnología",
    "Proceso",
    "Sub Proceso",
    "Servicios",
    "Precondiciones",
    "Descripción",
    "Paso a Paso",
    "Resultado Esperado",
    "Datos de Prueba",
]


def _safe_str(value, join_sep=", ") -> str:
    """Convierte cualquier valor a string de forma segura.
    Si es lista, une los elementos. Si es dict, lo serializa. Si es None, retorna ''."""
    if value is None:
        return ""
    if isinstance(value, list):
        return join_sep.join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


class AgenteMaestro:
    """Orquestador del sistema multi-agente con validación."""

    def __init__(self, kb: KnowledgeBase, llm_client: LMStudioClient = None, user_focus: str = ""):
        self.kb = kb
        self.llm = llm_client or get_llm_client()
        self.agente1 = Agente1Campos(kb, self.llm)
        self.agente2 = Agente2Detalle(kb, self.llm)
        self.casos_generados: list[dict] = []
        self.casos_descartados: int = 0
        self.user_focus = user_focus.strip() if user_focus else ""

    # ── 1. Planificación ────────────────────────────────────────────────

    def planificar_casos(self, json_req: dict, max_casos: int = None) -> list[dict]:
        """
        Planificación GUIADA POR LA MATRIZ.

        1. Extrae segmento y red del requerimiento
        2. Consulta la matriz para obtener TODAS las combos habilitadas
        3. Pide al LLM que PRIORICE y FILTRE las relevantes
        4. Agrega empaquetado (1Play, 2Play, 3Play) a cada combo
        """
        target = max_casos or agent_config.max_casos_por_requerimiento
        plan_target = min(target + max(3, target * 2 // 5), agent_config.max_casos_por_requerimiento * 2)

        contexto_parts = json_req.get("contexto", [])
        contexto = "\n".join(contexto_parts[:5]) if isinstance(contexto_parts, list) else str(contexto_parts)

        que_piden = json_req.get("que_piden", [])
        que_piden_texto = "\n".join(f"- {p}" for p in que_piden[:10]) if isinstance(que_piden, list) else str(que_piden)

        impacto = json_req.get("impacto_sistemas_bd", [])
        impacto_texto = "\n".join(f"- {i}" for i in impacto[:8]) if isinstance(impacto, list) else str(impacto)

        # ── Detectar segmento y red del requerimiento (para filtrar la matriz) ──
        segmento_detectado = self._detectar_segmento(json_req)
        red_detectada = self._detectar_red(json_req)

        # ── Obtener combinaciones habilitadas de la matriz ──
        combos_matriz = self.kb.get_combinaciones_habilitadas(segmento_detectado, red_detectada)

        # Aleatorizar orden para que el LLM no siempre elija las primeras
        # y así variar procesos, subprocesos y productos generados
        if combos_matriz:
            random.shuffle(combos_matriz)

        if combos_matriz:
            combos_lista = "\n".join(
                f"  {i+1}. {c['proceso']} → {c['subproceso']} [{c['control_origen']}]"
                for i, c in enumerate(combos_matriz)
            )
            print(f"  📊 Matriz: {len(combos_matriz)} combos habilitadas para {segmento_detectado or 'todos'}+{red_detectada or 'todas'}")
        else:
            combos_lista = "(No hay filtro de matriz disponible, usa tu criterio)"
            print(f"  ⚠️ Matriz: sin combos filtradas, se usará texto general")

        matrices_texto = self.kb.get_matrices_texto()

        # Sección de enfoque del usuario
        user_focus_section = ""
        if self.user_focus:
            user_focus_section = f"""
=== ENFOQUE SOLICITADO POR EL USUARIO ===
El usuario ha indicado: "{self.user_focus}"
DEBES priorizar combinaciones que se alineen con este enfoque."""

        prompt = f"""Eres un experto QA de telecomunicaciones de Claro/VTR Chile.
Analiza este requerimiento y SELECCIONA qué combinaciones de casos de prueba se deben generar.

=== REQUERIMIENTO ===
{contexto[:2000]}

Qué se pide:
{que_piden_texto[:1500]}

Impacto en sistemas:
{impacto_texto[:1000]}
{user_focus_section}

=== COMBINACIONES HABILITADAS POR LA MATRIZ (segmento={segmento_detectado or '?'}, red={red_detectada or '?'}) ===
{combos_lista}

=== MATRIZ COMPLETA DE REFERENCIA ===
{matrices_texto[:3000]}

=== EMPAQUETADOS DISPONIBLES ===
1Play (1 servicio), 2Play (2 servicios), 3Play (3 servicios), OTT, Prepago, PostPago, ClaroUP

=== MARCAS ===
Claro, VTR

=== REGLA DE MARCA POR CONTROL_ORIGEN ===
Cada subproceso tiene un código [CL/VT/Ambas] que indica para quién aplica:
- [CL]    → SOLO Claro → asignar marca: "Claro"
- [VT]    → SOLO VTR  → asignar marca: "VTR"
- [Ambas] → aplica a AMBAS marcas → puedes asignar "Claro" o "VTR"

=== INSTRUCCIONES ===
1. De las combinaciones habilitadas arriba, SELECCIONA las que aplican al requerimiento
2. Para cada combinación seleccionada, asigna:
   - marca: respetando SIEMPRE la regla de control_origen ([CL]=Claro, [VT]=VTR, [Ambas]=cualquiera)
   - empaquetado: 1Play, 2Play, 3Play, OTT, etc. (varía para cubrir más escenarios)
3. DEBES generar exactamente {plan_target} combinaciones
4. Cada combinación debe usar el sub_proceso EXACTO de la lista habilitada
5. Varía el empaquetado para que se prueben TODAS las configuraciones: mezcla 1Play, 2Play y 3Play de forma equitativa. No concentres todas las combinaciones en un solo empaquetado
6. Prioriza las combinaciones más relevantes al requerimiento
7. NO asignes marca "VTR" a un subproceso [CL], ni "Claro" a uno [VT]
8. DIVERSIDAD OBLIGATORIA: no repitas el mismo sub_proceso + empaquetado + marca más de 1 vez. Cubre el mayor número posible de sub_procesos distintos

IMPORTANTE: Tu respuesta debe ser ÚNICAMENTE un JSON válido.
Formato exacto:
{{"analisis": "breve análisis", "combinaciones": [{{"proceso": "Venta", "sub_proceso": "Venta Fija", "tecnologia": "{red_detectada or 'HFC'}", "marca": "Claro", "segmento": "{segmento_detectado or 'B2C'}", "empaquetado": "3Play"}}]}}"""

        print(f"🧠 Agente Maestro planificando casos (objetivo: {target}, planificando: {plan_target})...")

        combinaciones = []

        for attempt in range(2):
            try:
                kwargs = {
                    "temperature": agent_config.temperature_maestro,
                    "max_tokens": agent_config.max_tokens_maestro,
                }
                if attempt == 0:
                    kwargs["response_format"] = {"type": "json_object"}

                raw = self.llm.chat_with_retry(
                    [{"role": "user", "content": prompt}],
                    **kwargs,
                )

                parsed = JSONExtractor.extract(raw)
                combinaciones = parsed.get("combinaciones", [])
                analisis = parsed.get("analisis", "")
                if analisis:
                    print(f"  📊 Análisis: {analisis}")
                print(f"  📋 Combinaciones del LLM: {len(combinaciones)}")
                for c in combinaciones:
                    print(f"     → {c.get('proceso', '?')}/{c.get('sub_proceso', '?')}/{c.get('tecnologia', '?')}/{c.get('marca', '?')} [{c.get('empaquetado', '?')}]")
                break

            except Exception as e:
                sfx = " (sin response_format)" if attempt == 0 else ""
                print(f"  ⚠️ Intento {attempt + 1}{sfx} falló: {e}")
                if attempt == 0:
                    print(f"  🔄 Reintentando sin response_format...")
                else:
                    print(f"  📜 Respuesta raw (primeros 500 chars): {raw[:500] if 'raw' in dir() else 'N/A'}")

        # Fallback si LLM no devolvió nada
        if not combinaciones:
            print("  🔧 Usando fallback determinístico (basado en matriz)...")
            combinaciones = self._inferir_combinaciones_fallback(json_req)

        # ── Detectar marca y empaquetado del user_focus ──
        marcas_detectadas = self._detectar_marca(json_req)
        emp_detectado = self._detectar_empaquetado(json_req)

        # Asegurar que todas las combos tengan empaquetado y segmento
        for c in combinaciones:
            if not c.get("empaquetado"):
                c["empaquetado"] = emp_detectado
            if not c.get("segmento"):
                c["segmento"] = segmento_detectado or "B2C Residencial"

        # Expandir si el LLM propuso menos del objetivo
        if len(combinaciones) < plan_target:
            combinaciones = self._expandir_combinaciones(combinaciones, plan_target)

        return combinaciones

    def _detectar_segmento(self, json_req: dict) -> str:
        """Detecta el segmento del requerimiento analizando keywords."""
        texto = json.dumps(json_req, ensure_ascii=False).lower()
        if "b2b" in texto:
            if "pyme" in texto:
                return "B2B PYME"
            if "gran cuenta" in texto or "gc" in texto:
                return "B2B GC"
            return "B2B"
        if "b2c" in texto:
            if "pyme" in texto:
                return "B2C PYME"
            return "B2C Residencial"
        return "B2C Residencial"  # default

    def _detectar_red(self, json_req: dict) -> str:
        """Detecta la red/tecnología del requerimiento analizando keywords."""
        texto = json.dumps(json_req, ensure_ascii=False).lower()
        redes_detectadas = []
        if "neutra" in texto:
            redes_detectadas.append("NEUTRA")
        if "ftth" in texto or "fibra" in texto:
            redes_detectadas.append("FTTH")
        if "hfc" in texto:
            redes_detectadas.append("HFC")
        if "movil" in texto or "móvil" in texto or "celular" in texto:
            redes_detectadas.append("MOVIL")
        # Retornar la primera detectada, o None si no se detectó
        return redes_detectadas[0] if redes_detectadas else None

    def _detectar_marca(self, json_req: dict) -> list[str]:
        """Detecta la marca del user_focus y/o requerimiento.
        Returns lista de marcas a usar. Si el usuario explícitamente pidió una sola, retorna solo esa."""
        # Priorizar user_focus (instrucción directa del usuario)
        textos = []
        if self.user_focus:
            textos.append(self.user_focus.lower())
        textos.append(json.dumps(json_req, ensure_ascii=False).lower())

        for texto in textos:
            solo_claro = ("solo claro" in texto or "marca claro" in texto
                          or "todos claro" in texto or "deben ser marca claro" in texto
                          or "deben ser claro" in texto or "exclusivamente claro" in texto
                          or "únicamente claro" in texto or "unicamente claro" in texto)
            solo_vtr = ("solo vtr" in texto or "marca vtr" in texto
                        or "todos vtr" in texto or "deben ser marca vtr" in texto
                        or "deben ser vtr" in texto or "exclusivamente vtr" in texto)

            if solo_claro:
                return ["Claro"]
            if solo_vtr:
                return ["VTR"]

        return ["Claro", "VTR"]  # default: ambas

    def _detectar_empaquetado(self, json_req: dict) -> str:
        """Detecta el empaquetado del user_focus y/o requerimiento."""
        textos = []
        if self.user_focus:
            textos.append(self.user_focus.lower())
        textos.append(json.dumps(json_req, ensure_ascii=False).lower())

        for texto in textos:
            # Buscar "1 play", "1play", "1p"
            if "1 play" in texto or "1play" in texto or "servicio 1p" in texto or "1p " in texto:
                return "1Play"
            if "2 play" in texto or "2play" in texto or "servicio 2p" in texto or "2p " in texto:
                return "2Play"
            if "3 play" in texto or "3play" in texto or "servicio 3p" in texto or "3p " in texto:
                return "3Play"
            if "ott" in texto:
                return "OTT"
            if "prepago" in texto:
                return "Prepago"
            if "postpago" in texto:
                return "PostPago"
            if "claroup" in texto:
                return "ClaroUP"

        return "3Play"  # default

    def _inferir_combinaciones_fallback(self, json_req: dict) -> list[dict]:
        """Infiere combinaciones usando la MATRIZ cuando el LLM falla.
        Respeta user_focus para marca y empaquetado."""
        segmento = self._detectar_segmento(json_req)
        red = self._detectar_red(json_req)
        marcas = self._detectar_marca(json_req)
        empaquetado = self._detectar_empaquetado(json_req)

        print(f"  📌 Restricciones detectadas: marcas={marcas}, empaquetado={empaquetado}, segmento={segmento}, red={red}")

        # Obtener combos habilitadas de la matriz
        combos_matriz = self.kb.get_combinaciones_habilitadas(segmento, red)

        if combos_matriz:
            # Aleatorizar para variar productos entre ejecuciones
            random.shuffle(combos_matriz)
            combinaciones = []
            for c in combos_matriz:
                ctrl = c.get("control_origen", "Ambas")
                for marca in marcas:
                    # Respetar restricción de marca por control_origen
                    if ctrl == "CL" and marca != "Claro":
                        continue
                    if ctrl == "VT" and marca != "VTR":
                        continue
                    combinaciones.append({
                        "proceso": c["proceso"],
                        "sub_proceso": c["subproceso"],
                        "tecnologia": red or "HFC",
                        "marca": marca,
                        "segmento": segmento,
                        "empaquetado": empaquetado,
                        "control_origen": ctrl,
                    })
            print(f"  📋 Combinaciones desde matriz (fallback): {len(combinaciones)}")
            for c in combinaciones[:10]:
                print(f"     → {c['proceso']}/{c['sub_proceso']}/{c['tecnologia']}/{c['marca']} [{c['empaquetado']}]")
            if len(combinaciones) > 10:
                print(f"     ... y {len(combinaciones)-10} más")
            return combinaciones

        # Fallback extremo si no hay matriz
        marcas = self._detectar_marca(json_req)
        empaquetado = self._detectar_empaquetado(json_req)
        texto = json.dumps(json_req, ensure_ascii=False).lower()
        keyword_proceso = {
            "venta": ("Venta", "Venta Fija"),
            "alta": ("Venta", "Crear Cliente"),
            "reclamo": ("Postventa", "Reparaciones"),
            "desconexión": ("Postventa", "Desconexiones"),
            "desconexion": ("Postventa", "Desconexiones"),
            "reparación": ("Postventa", "Reparaciones"),
            "reparacion": ("Postventa", "Reparaciones"),
            "facturación": ("Facturacion", "Proceso de Facturación (emisión de boletas / facturas)"),
            "facturacion": ("Facturacion", "Proceso de Facturación (emisión de boletas / facturas)"),
            "cambio de plan": ("Postventa", "Cambio de Promo (1 a 2, 1 a 3, 2 a 3)"),
            "suspensión": ("Postventa", "Suspensión Voluntaria"),
            "suspension": ("Postventa", "Suspensión Voluntaria"),
            "traslado": ("Postventa", "Traslado de Domicilio"),
        }

        procesos_detectados = set()
        for keyword, (proceso, sub) in keyword_proceso.items():
            if keyword in texto:
                procesos_detectados.add((proceso, sub))

        if not procesos_detectados:
            procesos_detectados = {("Venta", "Venta Fija")}

        combinaciones = []
        for proc, sub in sorted(procesos_detectados):
            for marca in marcas:
                combinaciones.append({
                    "proceso": proc,
                    "sub_proceso": sub,
                    "tecnologia": red or "HFC",
                    "marca": marca,
                    "segmento": segmento,
                    "empaquetado": empaquetado,
                })

        print(f"  📋 Combinaciones inferidas (fallback keywords): {len(combinaciones)}")
        for c in combinaciones:
            print(f"     → {c['proceso']}/{c['sub_proceso']}/{c['tecnologia']}/{c['marca']} [{c['empaquetado']}]")
        return combinaciones

    def _expandir_combinaciones(self, combinaciones: list[dict], target: int) -> list[dict]:
        """
        Expande la lista de combinaciones hasta alcanzar el target.

        Estrategia de expansión (en orden de prioridad):
        1. Variar empaquetado (1Play → 2Play → 3Play)
        2. Variar marca (Claro ↔ VTR)
        3. Si aún faltan, ciclar las existentes
        """
        if len(combinaciones) >= target:
            return combinaciones[:target]

        # Respetar restricciones del user_focus
        marcas = self._detectar_marca({})  # solo analiza user_focus
        emp_detectado = self._detectar_empaquetado({})  # solo analiza user_focus
        # Si el usuario fijó un empaquetado específico, no variar
        empaquetados_para_expandir = [emp_detectado] if emp_detectado != "3Play" else EMPAQUETADOS

        # Indexar combinaciones existentes para evitar duplicados
        existentes = set()
        for c in combinaciones:
            key = (c.get("proceso", ""), c.get("sub_proceso", ""), c.get("tecnologia", ""),
                   c.get("marca", ""), c.get("empaquetado", ""))
            existentes.add(key)

        expandidas = list(combinaciones)

        # Ronda 1: Variar empaquetado
        for c in list(combinaciones):
            if len(expandidas) >= target:
                break
            for emp in empaquetados_para_expandir:
                if len(expandidas) >= target:
                    break
                key = (c.get("proceso", ""), c.get("sub_proceso", ""), c.get("tecnologia", ""),
                       c.get("marca", ""), emp)
                if key not in existentes:
                    new = dict(c)
                    new["empaquetado"] = emp
                    expandidas.append(new)
                    existentes.add(key)

        # Ronda 2: Variar marca
        for c in list(combinaciones):
            if len(expandidas) >= target:
                break
            for marca in marcas:
                if len(expandidas) >= target:
                    break
                key = (c.get("proceso", ""), c.get("sub_proceso", ""), c.get("tecnologia", ""),
                       marca, c.get("empaquetado", ""))
                if key not in existentes:
                    new = dict(c)
                    new["marca"] = marca
                    expandidas.append(new)
                    existentes.add(key)

        # Ronda 3: Variar empaquetado + marca cruzado
        for c in list(combinaciones):
            if len(expandidas) >= target:
                break
            for emp in empaquetados_para_expandir:
                for marca in marcas:
                    if len(expandidas) >= target:
                        break
                    key = (c.get("proceso", ""), c.get("sub_proceso", ""), c.get("tecnologia", ""),
                           marca, emp)
                    if key not in existentes:
                        new = dict(c)
                        new["empaquetado"] = emp
                        new["marca"] = marca
                        expandidas.append(new)
                        existentes.add(key)

        # Ronda 4: Si aún faltan, ciclar las existentes
        if len(expandidas) < target:
            base = list(expandidas)
            idx = 0
            while len(expandidas) < target:
                expandidas.append(dict(base[idx % len(base)]))
                idx += 1

        added = len(expandidas) - len(combinaciones)
        if added > 0:
            print(f"  📈 Expandidas +{added} combinaciones para alcanzar objetivo ({len(expandidas)} total)")
            for c in expandidas[len(combinaciones):]:
                print(f"     + {c.get('proceso', '?')}/{c.get('sub_proceso', '?')}/{c.get('tecnologia', '?')}/{c.get('marca', '?')} [{c.get('empaquetado', '?')}]")

        return expandidas

    # ── 2. Validación de Cabecera (Agente 1) ────────────────────────────

    def validar_cabecera(self, campos: dict, combinacion: dict) -> dict:
        """
        Validación DETERMINÍSTICA de la cabecera del Agente 1.
        Incluye validación contra la matriz.
        Returns: {"valido": bool, "errores": list[str]}
        """
        errores = []

        # Validar Proceso
        proceso = _safe_str(campos.get("Proceso", "")).strip()
        if not proceso:
            errores.append("Proceso está vacío")

        # Validar Tecnología
        tec = _safe_str(campos.get("Tecnología", "")).strip()
        tec_permitidas = ["FTTH", "HFC", "NEUTRA", "NFTT", "CFTT", "MOVIL", "SIP", "XGSPON", "GPON", "N/A"]
        if tec and tec.upper() not in [t.upper() for t in tec_permitidas]:
            errores.append(f"Tecnología '{tec}' no está en valores permitidos: {tec_permitidas}")

        # Validar Marca
        marca = _safe_str(campos.get("Marca", "")).strip()
        if marca and marca not in ["Claro", "VTR", "CLARO", "claro", "vtr"]:
            errores.append(f"Marca '{marca}' no es válida (debe ser Claro o VTR)")

        # Validar Sub Proceso
        sub = _safe_str(campos.get("Sub Proceso", "")).strip()
        if not sub:
            errores.append("Sub Proceso está vacío")

        # Validar contra la matriz (proceso + sub_proceso + segmento + red)
        if proceso and sub and tec:
            segmento = _safe_str(campos.get("Segmento", "")).strip()
            val_matriz = self.kb.validar_combinacion(proceso, sub, segmento, tec)
            if not val_matriz.get("valido", True):
                errores.append(f"Matriz: {val_matriz['razon']}")

        return {"valido": len(errores) == 0, "errores": errores}

    # ── 3. Validación de Detalle (Agente 2) ─────────────────────────────

    def validar_detalle(self, detalle: dict, cabecera: dict) -> dict:
        """
        Validación HÍBRIDA del detalle del Agente 2.
        - Determinística: formato y plataformas
        - LLM: coherencia con cabecera
        Returns: {"valido": bool, "errores": list[str]}
        """
        errores = []

        # --- Validaciones determinísticas ---

        # Verificar que Paso a Paso no esté vacío
        pasos = _safe_str(detalle.get("Paso a Paso", ""), join_sep="\n").strip()
        if not pasos or len(pasos) < 30:
            errores.append("Paso a Paso está vacío o es demasiado corto")

        # Verificar estructura (secciones con romanos)
        tiene_romanos = bool(re.search(r"^[IVX]+\.", pasos, re.MULTILINE))
        tiene_numerados = bool(re.search(r"^\d+\.", pasos, re.MULTILINE))
        if not tiene_romanos and not tiene_numerados:
            errores.append(
                "Los Pasos no siguen la estructura Mantis. "
                "Deben tener secciones con números romanos (I., II.) "
                "y pasos con números arábigos (1., 2.)"
            )

        # Verificar que mencione al menos 1 plataforma real de Siebel
        sistemas_mencionados = []
        for sistema in self.kb.siebel_sistemas:
            # Buscar el nombre del sistema (case-insensitive)
            if sistema.lower() in pasos.lower():
                sistemas_mencionados.append(sistema)

        if len(sistemas_mencionados) < 1:
            errores.append(
                f"Los Pasos no mencionan ninguna plataforma real del mapa Siebel. "
                f"Disponibles: {', '.join(self.kb.siebel_sistemas[:10])}"
            )

        # Verificar Descripción (ahora generada por Agente 2)
        desc = _safe_str(detalle.get("Descripción", "")).strip()
        if not desc or len(desc) < 10:
            errores.append("Descripción está vacía o es demasiado corta (min 10 caracteres)")

        # Verificar Precondiciones
        precon = _safe_str(detalle.get("Precondiciones", "")).strip()
        if not precon or len(precon) < 5:
            errores.append("Precondiciones está vacío o es demasiado corto")

        # Verificar Resultado Esperado
        resultado = _safe_str(detalle.get("Resultado Esperado", "")).strip()
        if not resultado or len(resultado) < 20:
            errores.append("Resultado Esperado está vacío o es demasiado corto")

        # Verificar formato de resultado (debería tener * o -)
        if resultado and not any(c in resultado for c in ["*", "-", "•"]):
            errores.append(
                "Resultado Esperado debe usar formato de lista con * al inicio de cada validación"
            )

        # Si los errores determinísticos son graves, no gastar tokens en LLM
        if errores:
            return {"valido": False, "errores": errores}

        # --- Validación LLM: coherencia ---
        coherencia = self._validar_coherencia_llm(detalle, cabecera)
        if not coherencia.get("valido", True):
            errores.extend(coherencia.get("errores", []))

        return {"valido": len(errores) == 0, "errores": errores}

    def _validar_coherencia_llm(self, detalle: dict, cabecera: dict) -> dict:
        """Usa LLM para validar coherencia entre detalle y cabecera."""
        # Obtener ejemplo real similar para comparar
        flujo = cabecera.get("Proceso", "")
        marca = cabecera.get("Marca", "Claro")
        ejemplos = self.kb.get_mantis_ejemplo(flujo, marca, 1)
        ejemplo_texto = ""
        if ejemplos:
            ejemplo_texto = self.kb.formatear_mantis_ejemplo(ejemplos[0])

        prompt = f"""Eres un validador QA experto de Claro/VTR Chile.
Valida si el detalle del caso de prueba es COHERENTE con la cabecera.

CABECERA:
Proceso: {cabecera.get('Proceso', '')}
Sub Proceso: {cabecera.get('Sub Proceso', '')}
Tecnología: {cabecera.get('Tecnología', '')}
Marca: {cabecera.get('Marca', '')}

DETALLE A VALIDAR:
Precondiciones: {detalle.get('Precondiciones', '')}
Descripción: {detalle.get('Descripción', '')}
Paso a Paso: {detalle.get('Paso a Paso', '')[:1500]}
Resultado Esperado: {detalle.get('Resultado Esperado', '')}

EJEMPLO REAL SIMILAR:
{ejemplo_texto[:1000]}

SISTEMAS REALES: {', '.join(self.kb.siebel_sistemas)}

Valida:
1. ¿Los pasos son coherentes con el proceso "{flujo}"?
2. ¿El resultado esperado corresponde a los pasos?
3. ¿Se mencionan plataformas reales?

Responde SOLO JSON:
{{"valido": true/false, "errores": ["error1", "error2"]}}"""

        try:
            raw = self.llm.chat_with_retry(
                [{"role": "user", "content": prompt}],
                temperature=0.0,
                max_tokens=500,
            )
            return JSONExtractor.extract(raw)
        except Exception:
            # Si falla la validación LLM, asumir coherente
            return {"valido": True, "errores": []}

    # ── 4. Ejecución principal ──────────────────────────────────────────

    def ejecutar(
        self,
        json_req: dict,
        max_casos: int = None,
        output_dir: Path = None,
        cancel_event=None,
    ) -> list[dict]:
        """
        Ejecuta el ciclo completo de generación multi-agente con reposición.

        1. Planifica combinaciones
        2. Para cada combinación: Agente 1 → valida → Agente 2 → valida
        3. Guarda CSV/Excel INCREMENTALMENTE después de cada caso validado
        4. Si no se alcanza max_casos, genera rondas de reposición (hasta 3 extra)
        """
        max_casos = max_casos or agent_config.max_casos_por_requerimiento
        max_retries = agent_config.max_retries_validacion
        MAX_RONDAS_REPOSICION = 3

        # Preparar rutas de output incremental
        if output_dir:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            self._csv_path = output_dir / "casos_prueba.csv"
            self._xlsx_path = output_dir / "casos_prueba.xlsx"
        else:
            self._csv_path = None
            self._xlsx_path = None

        # Paso 1: Planificar (pedir más combinaciones para compensar descartes)
        combinaciones = self.planificar_casos(json_req, max_casos=max_casos)

        print(f"\n{'='*60}")
        print(f"🚀 Generando hasta {max_casos} casos de prueba (planificadas: {len(combinaciones)} combinaciones)...")
        if self._csv_path:
            print(f"💾 Guardado incremental en: {output_dir}")
        print(f"{'='*60}\n")

        errores_runtime = []  # Acumula errores no-fatales para el resumen
        caso_global_idx = 0   # Contador global de intentos
        ronda_actual = 0
        cancelado = False
        combinaciones_base = list(combinaciones[:6])  # Guardar primeras combos como semilla

        # ── Ciclo de rondas (principal + reposición) ──
        while len(self.casos_generados) < max_casos:
            ronda_actual += 1

            if ronda_actual == 1:
                lote = combinaciones
            else:
                # Rondas de reposición
                if ronda_actual - 1 > MAX_RONDAS_REPOSICION:
                    print(f"\n⚠️ Se alcanzó el límite de {MAX_RONDAS_REPOSICION} rondas de reposición.")
                    print(f"   Casos generados: {len(self.casos_generados)}/{max_casos}")
                    break

                faltantes = max_casos - len(self.casos_generados)
                # Agregar margen extra del 50% para compensar posibles nuevos descartes
                margen = max(2, faltantes // 2)
                lote_size = faltantes + margen

                print(f"\n{'='*60}")
                print(f"🔄 Ronda de reposición {ronda_actual - 1}/{MAX_RONDAS_REPOSICION}: "
                      f"faltan {faltantes} casos, generando {lote_size} combinaciones extra...")
                print(f"{'='*60}\n")

                # Generar nuevas combinaciones variando las semillas base
                lote = self._expandir_combinaciones(combinaciones_base, lote_size)

            for combo in lote:
                if cancel_event and cancel_event.is_set():
                    print(f"\n⚠️ Proceso cancelado por el usuario. "
                          f"Casos generados hasta ahora: {len(self.casos_generados)}")
                    cancelado = True
                    break

                if len(self.casos_generados) >= max_casos:
                    print(f"⚡ Límite de {max_casos} casos alcanzado.")
                    break

                caso_global_idx += 1

                try:
                    print(f"\n{'─'*40}")
                    print(f"📝 Caso {caso_global_idx} "
                          f"(generados: {len(self.casos_generados)}/{max_casos}): "
                          f"{combo.get('proceso')}/{combo.get('sub_proceso', '?')}/{combo.get('tecnologia')}/{combo.get('marca')}")
                    print(f"{'─'*40}")

                    # ── Fase 1: Agente 1 (Cabecera) ──
                    cabecera = None
                    errores_a1 = None

                    for intento in range(max_retries):
                        cabecera_raw = self.agente1.generar(json_req, combo, errores_a1, user_focus=self.user_focus)
                        validacion = self.validar_cabecera(cabecera_raw, combo)

                        if validacion["valido"]:
                            cabecera = cabecera_raw
                            print(f"  ✅ Cabecera validada (intento {intento + 1})")
                            break
                        else:
                            errores_a1 = validacion["errores"]
                            print(f"  ❌ Cabecera rechazada (intento {intento + 1}): {errores_a1}")

                    if not cabecera:
                        print(f"  ⚠️ Caso descartado: cabecera no válida tras {max_retries} intentos")
                        self.casos_descartados += 1
                        continue

                    # ── Fase 2: Agente 2 (Detalle) ──
                    detalle = None
                    errores_a2 = None

                    for intento in range(max_retries):
                        detalle_raw = self.agente2.generar(json_req, cabecera, errores_a2, user_focus=self.user_focus)
                        validacion = self.validar_detalle(detalle_raw, cabecera)

                        if validacion["valido"]:
                            detalle = detalle_raw
                            print(f"  ✅ Detalle validado (intento {intento + 1})")
                            break
                        else:
                            errores_a2 = validacion["errores"]
                            print(f"  ❌ Detalle rechazado (intento {intento + 1}): {errores_a2}")

                    if not detalle:
                        print(f"  ⚠️ Caso descartado: detalle no válido tras {max_retries} intentos")
                        self.casos_descartados += 1
                        continue

                    # ── Ensamblar caso completo ──
                    caso_completo = self._ensamblar_caso(len(self.casos_generados) + 1, cabecera, detalle)
                    self.casos_generados.append(caso_completo)
                    print(f"  🎯 Caso #{len(self.casos_generados)} completado y guardado")

                    # ── Guardar incrementalmente ──
                    self._guardar_incremental()

                except Exception as e:
                    error_msg = (f"Error en caso {caso_global_idx} "
                                 f"({combo.get('proceso', '?')}): {type(e).__name__}: {e}")
                    errores_runtime.append(error_msg)
                    print(f"  ⚠️ {error_msg} — saltando al siguiente caso")
                    self.casos_descartados += 1
                    continue

            # Si se canceló o ya se alcanzó el objetivo, salir del while
            if cancelado or len(self.casos_generados) >= max_casos:
                break

        # ── Resumen final ──
        print(f"\n{'='*60}")
        if len(self.casos_generados) >= max_casos:
            print(f"✅ Generación completada: {len(self.casos_generados)} casos generados "
                  f"(objetivo: {max_casos} ✔️), {self.casos_descartados} descartados")
        else:
            print(f"⚠️ Generación finalizada: {len(self.casos_generados)}/{max_casos} casos generados, "
                  f"{self.casos_descartados} descartados")
        if ronda_actual > 1:
            print(f"🔄 Se realizaron {ronda_actual - 1} ronda(s) de reposición")
        if errores_runtime:
            print(f"⚠️ {len(errores_runtime)} error(es) no-fatales durante la generación:")
            for err in errores_runtime:
                print(f"   - {err}")
        if self._csv_path:
            print(f"💾 Archivos finales:")
            print(f"   CSV:  {self._csv_path}")
            if HAS_OPENPYXL:
                print(f"   XLSX: {self._xlsx_path}")
        print(f"{'='*60}")

        return self.casos_generados

    def _ensamblar_caso(self, id_caso: int, cabecera: dict, detalle: dict) -> dict:
        """Ensambla un caso completo con las 14 columnas del CSV final."""
        return {
            "ID": id_caso,
            "Tipo de Prueba": _safe_str(cabecera.get("Tipo de Prueba", "Proyecto (Funcional)")),
            "Prioridad": cabecera.get("Prioridad", 1),
            "Marca": _safe_str(cabecera.get("Marca", "")),
            "Segmento": _safe_str(cabecera.get("Segmento", "B2C Residencial")),
            "Tecnología": _safe_str(cabecera.get("Tecnología", "")),
            "Proceso": _safe_str(cabecera.get("Proceso", "")),
            "Sub Proceso": _safe_str(cabecera.get("Sub Proceso", "")),
            "Servicios": _safe_str(cabecera.get("Servicios", "")),
            "Precondiciones": _safe_str(detalle.get("Precondiciones", "N/A")),
            "Descripción": _safe_str(detalle.get("Descripción", "")),
            "Paso a Paso": _safe_str(detalle.get("Paso a Paso", ""), join_sep="\n"),
            "Resultado Esperado": _safe_str(detalle.get("Resultado Esperado", ""), join_sep="\n"),
            "Datos de Prueba": _safe_str(detalle.get("Datos de Prueba", "")),
        }

    # ── 5. Exportación ──────────────────────────────────────────────────

    def _guardar_incremental(self):
        """Guarda CSV y Excel en cada iteración para no perder progreso."""
        if not self._csv_path:
            return
        try:
            self._escribir_csv(self._csv_path)
        except Exception as e:
            print(f"  ⚠️ Error escribiendo CSV: {e}")
        if HAS_OPENPYXL:
            self._escribir_xlsx(self._xlsx_path)
        n = len(self.casos_generados)
        print(f"  💾 Guardado incremental: {n} caso{'s' if n > 1 else ''} en disco")

    def _escribir_csv(self, path: Path, delimiter: str = ","):
        """Escribe todos los casos al CSV (sobrescribiendo)."""
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=COLUMNAS_CSV, delimiter=delimiter,
                extrasaction="ignore",
                quoting=csv.QUOTE_ALL,
            )
            writer.writeheader()
            for caso in self.casos_generados:
                # Reemplazar newlines reales por \n literal para que cada caso
                # quede en UNA sola línea del CSV
                caso_limpio = {}
                for k, v in caso.items():
                    if isinstance(v, str):
                        caso_limpio[k] = v.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")
                    else:
                        caso_limpio[k] = v
                writer.writerow(caso_limpio)

    def _escribir_xlsx(self, path: Path):
        """Escribe todos los casos a Excel (sobrescribiendo)."""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Casos de Prueba"

            # Header
            for col_idx, col_name in enumerate(COLUMNAS_CSV, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = openpyxl.styles.Font(bold=True)

            # Data — convertir cada valor a string seguro para evitar errores de inserción
            for row_idx, caso in enumerate(self.casos_generados, 2):
                for col_idx, col_name in enumerate(COLUMNAS_CSV, 1):
                    valor = caso.get(col_name, "")
                    # openpyxl no acepta listas/dicts — convertir a string
                    if isinstance(valor, (list, dict)):
                        valor = _safe_str(valor)
                    # Truncar strings muy largos (Excel tiene límite de 32767 chars por celda)
                    if isinstance(valor, str) and len(valor) > 32000:
                        valor = valor[:32000] + "... [truncado]"
                    ws.cell(row=row_idx, column=col_idx, value=valor)

            # Autofit column widths (approximate)
            for col_idx, col_name in enumerate(COLUMNAS_CSV, 1):
                max_len = len(col_name)
                for row_idx in range(2, len(self.casos_generados) + 2):
                    val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    max_len = max(max_len, min(len(val), 60))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 65)

            wb.save(path)
        except Exception as e:
            print(f"  ⚠️ Error escribiendo Excel: {e} — el CSV sigue disponible")

    def exportar_csv(self, output_path: Path, delimiter: str = ",") -> Path:
        """Exporta los casos generados a CSV con formato Mantis."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._escribir_csv(output_path, delimiter)
        print(f"📄 CSV exportado: {output_path} ({len(self.casos_generados)} casos)")
        return output_path

    def exportar_xlsx(self, output_path: Path) -> Path:
        """Exporta los casos generados a Excel."""
        if not HAS_OPENPYXL:
            print("⚠️ openpyxl no instalado, Excel no generado")
            return None
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._escribir_xlsx(output_path)
        print(f"📊 Excel exportado: {output_path} ({len(self.casos_generados)} casos)")
        return output_path

    def get_resumen(self) -> str:
        """Genera un resumen textual de los casos generados."""
        lines = [
            f"# Resumen de Generación de Casos de Prueba",
            f"- Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"- Casos generados: {len(self.casos_generados)}",
            f"- Casos descartados: {self.casos_descartados}",
            "",
            "## Casos Generados:",
        ]
        for caso in self.casos_generados:
            lines.append(
                f"- [{caso['ID']}] {caso['Proceso']} / {caso['Tecnología']} / "
                f"{caso['Marca']}: {caso['Descripción'][:80]}"
            )
        return "\n".join(lines)
